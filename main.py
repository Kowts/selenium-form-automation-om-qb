import os
import time
import logging
import pandas as pd
from dotenv import load_dotenv
from selenium import webdriver
from openpyxl import load_workbook
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import ElementNotInteractableException, NoSuchElementException

# Carregar variáveis de ambiente do ficheiro .env
load_dotenv()

# Configurar logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# Obter email e password das variáveis de ambiente
EMAIL = os.getenv('LOGIN_EMAIL')
PASSWORD = os.getenv('LOGIN_PASSWORD')
URL = os.getenv('LOGIN_URL')

# Garantir que email, password e URL não sejam None
if not EMAIL or not PASSWORD or not URL:
    logging.error("Email, password ou URL não encontrados nas variáveis de ambiente.")
    raise ValueError("Faltando LOGIN_EMAIL, LOGIN_PASSWORD ou LOGIN_URL no ficheiro .env")

# Versão do Chrome - especificar uma versão ou usar a mais recente
chrome_version = "129.0.6668.90"

# Caminho do ficheiro
excel_file_path = "ficheiro/BB.xlsx"

# Carregar dados do Excel
try:
    logging.info(f"Carregando ficheiro Excel de: {excel_file_path}")
    excel_data = pd.read_excel(excel_file_path)
    logging.info("Dados do Excel carregados com sucesso.")
except Exception as e:
    logging.error(f"Falha ao carregar o ficheiro Excel: {e}")
    raise

# Configurar o WebDriver para o Chrome
try:
    logging.info("Configurando WebDriver do Chrome...")
    service = Service(ChromeDriverManager(driver_version=chrome_version).install())
    driver = webdriver.Chrome(service=service)
    logging.info("WebDriver configurado com sucesso.")
except Exception as e:
    logging.error(f"Falha ao inicializar o WebDriver do Chrome: {e}")
    raise

# Navegar para a página de login
driver.get(URL)

# Login no website
try:
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="id_login"]')))
    logging.info("Fazendo login...")
    driver.find_element(By.XPATH, '//*[@id="id_login"]').send_keys(EMAIL)
    driver.find_element(By.XPATH, '//*[@id="id_password"]').send_keys(PASSWORD)
    driver.find_element(By.XPATH, '//*[@id="auth-right"]/form/button').click()
    logging.info("Login enviado.")
except Exception as e:
    logging.error(f"Falha ao fazer login: {e}")
    driver.quit()
    raise

# Navegar para a página do menu
try:
    WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="sidebar"]/ul/li[3]/a')))
    menu_link = driver.find_element(By.XPATH, '//*[@id="sidebar"]/ul/li[3]/a')
    driver.execute_script("arguments[0].click();", menu_link)  # Usando JavaScript para clicar

    submenu_link = driver.find_element(By.XPATH, '//*[@id="sidebar"]/ul/li[3]/ul/li[1]/a')
    driver.execute_script("arguments[0].click();", submenu_link)  # Usando JavaScript para clicar

    logging.info("Navegado para a página do menu.")
except Exception as e:
    logging.error(f"Falha ao navegar para a página do menu: {e}")
    driver.quit()
    raise

# Abrir o excel uma vez antes do loop
wb = load_workbook(excel_file_path)
ws = wb.active

# Loop pelas linhas do Excel e preencher os campos do formulário no website
for i, row in excel_data.iterrows():
    retries = 3  # Número de tentativas de reenvio caso haja falha
    for attempt in range(retries):
        try:
            # Obter os dados do formulário da linha do Excel
            BH = str(row['@BH'])
            BD = str(row['@BD'])
            BA = str(row['@BA'])

            # Determinar se o formulário está no estado de pré-envio ou pós-envio
            try:
                driver.find_element(By.XPATH, '/html/body/div[4]/div[1]/div[2]/div/div/form/div[2]/div[1]/input')
                submission_state = "before"
            except NoSuchElementException:
                submission_state = "after"

            # Definir o XPath correto para os campos com base no estado de envio
            if submission_state == "before":
                xpaths = [
                    '/html/body/div[4]/div[1]/div[2]/div/div/form/div[2]/div[1]/input',
                    '/html/body/div[4]/div[1]/div[2]/div/div/form/div[2]/div[2]/input',
                    '/html/body/div[4]/div[1]/div[2]/div/div/form/div[2]/div[3]/input',
                    '/html/body/div[4]/div[1]/div[2]/div/div/form/div[3]/button'
                ]
            else:  # Após o envio do formulário
                xpaths = [
                    '/html/body/div[4]/div[1]/div[3]/div/div/form/div[2]/div[1]/input',
                    '/html/body/div[4]/div[1]/div[3]/div/div/form/div[2]/div[2]/input',
                    '/html/body/div[4]/div[1]/div[3]/div/div/form/div[2]/div[3]/input',
                    '/html/body/div[4]/div[1]/div[3]/div/div/form/div[3]/button'
                ]

            # Esperar o campo de entrada estar clicável e inserir os dados
            WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, xpaths[0]))).send_keys(BH)
            driver.find_element(By.XPATH, xpaths[1]).send_keys(BD)
            driver.find_element(By.XPATH, xpaths[2]).send_keys(BA)

            # Enviar o formulário
            driver.find_element(By.XPATH, xpaths[3]).click()

            logging.info(f"Formulário enviado para a linha {i + 1}: BH={BH}, BD={BD}, BA={BA}")

            # Recuperar a mensagem de alerta após o envio do formulário
            try:
                alert_message_element = WebDriverWait(driver, 10).until(
                    EC.visibility_of_element_located((By.XPATH, '/html/body/div[4]/div[1]/div[2]/strong'))
                )
                alert_message = alert_message_element.text
            except NoSuchElementException:
                logging.error("Mensagem de alerta não encontrada")
                alert_message = None

            if alert_message:
                # Extrair a parte da mensagem após ":"
                if ":" in alert_message:
                    alert_message = alert_message.split(":", 1)[1].strip()  # Dividir e pegar a parte após o primeiro ":"

                logging.info(f"Mensagem de alerta para a linha {i + 1}: {alert_message}")

                # Escrever a mensagem de alerta na coluna G
                ws[f'G{i + 2}'] = alert_message

                # Guardar o excel após atualizar a mensagem de alerta na coluna G
                wb.save(excel_file_path)

            # Quebrar o loop de tentativas após o sucesso
            time.sleep(1)  # Esperar após o envio antes de passar para a próxima linha
            break

        except ElementNotInteractableException as e:
            logging.error(f"Tentativa {attempt + 1} - Elemento não interativo para a linha {i + 1}: {e}")
            time.sleep(2)  # Esperar antes de tentar novamente
        except Exception as e:
            logging.error(f"Falha ao enviar o formulário para a linha {i + 1}: {e}")
            continue

# Guardar o excel após todas as linhas terem sido processadas
#wb.save(excel_file_path)
logging.info("Navegador fechado. Script concluído.")
